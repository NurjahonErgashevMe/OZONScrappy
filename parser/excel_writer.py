import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelWriter:
    def __init__(self):
        self.logger = logging.getLogger('excel_writer')
        self.output_dir = "output"
        
        # Создаем папку output если её нет
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def save_to_excel(self, products, seller_name):
        """Сохранение товаров в Excel файл"""
        try:
            # Создаем имя файла с timestamp
            timestamp = datetime.now().strftime("%d.%m.%Y_%H-%M-%S")
            filename = f"{seller_name}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            # Создаем рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Товары"
            
            # Заголовки
            headers = [
                "Название товара",
                "Цена со скидкой",
                "Цена без скидки", 
                "Скидка (%)",
                "Рейтинг",
                "Количество отзывов",
                "Ссылка на товар"
            ]
            
            # Настройка заголовков
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self._get_border()
            
            # Заполняем данными
            for row, product in enumerate(products, 2):
                ws.cell(row=row, column=1, value=product['name'])
                ws.cell(row=row, column=2, value=product['price_with_discount'])
                ws.cell(row=row, column=3, value=product['price_without_discount'])
                ws.cell(row=row, column=4, value=product['discount_percent'])
                ws.cell(row=row, column=5, value=product['rating'])
                ws.cell(row=row, column=6, value=product['reviews_count'])
                ws.cell(row=row, column=7, value=product['url'])
                
                # Применяем стили к ячейкам
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self._get_border()
                    cell.alignment = Alignment(vertical="center")
                    
                    # Особые стили для разных колонок
                    if col in [2, 3, 4]:  # Цены и скидка
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col in [5, 6]:  # Рейтинг и отзывы
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Настройка ширины колонок
            column_widths = {
                'A': 75,  # Название
                'B': 40,  # Цена со скидкой
                'C': 40,  # Цена без скидки
                'D': 40,  # Скидка
                'E': 40,  # Рейтинг
                'F': 40,  # Отзывы
                'G': 75   # Ссылка
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Настройка высоты строк
            for row in range(1, len(products) + 2):
                ws.row_dimensions[row].height = 25
            
            # Применяем автофильтр
            ws.auto_filter.ref = f"A1:G{len(products) + 1}"
            
            # Замораживаем первую строку
            ws.freeze_panes = "A2"
            
            # Сохраняем файл
            wb.save(filepath)
            self.logger.info(f"Excel файл сохранен: {filepath}")
            
            return filepath
            
        except Exception as e:
            self.logger.error(f"Ошибка сохранения Excel файла: {str(e)}")
            return None
    
    def save_sellers_to_excel(self, sellers_data, filename_prefix):
        """Сохранение данных о продавцах в Excel файл"""
        try:
            # Создаем имя файла с timestamp
            timestamp = datetime.now().strftime("%d.%m.%Y_%H-%M-%S")
            filename = f"{self._clean_filename(filename_prefix)}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            # Создаем рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Продавцы ИНН"
            
            # Заголовки для данных о продавцах
            headers = [
                "Название продавца",
                "Название компании",
                "ИНН",
                "Ссылка на продавца"
            ]
            
            # Настройка заголовков
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self._get_border()
            
            # Заполняем данными о продавцах
            for row, seller in enumerate(sellers_data, 2):
                ws.cell(row=row, column=1, value=seller['name'])  # Название продавца
                ws.cell(row=row, column=2, value=seller['price_with_discount'])  # Название компании
                ws.cell(row=row, column=3, value=seller['price_without_discount'])  # ИНН
                ws.cell(row=row, column=4, value=seller['url'])  # Ссылка
                
                # Применяем стили к ячейкам
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self._get_border()
                    cell.alignment = Alignment(vertical="center")
                    
                    # Выравнивание для ИНН по центру
                    if col == 3:  # ИНН
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Настройка ширины колонок для продавцов
            column_widths = {
                'A': 50,  # Название продавца
                'B': 60,  # Название компании
                'C': 20,  # ИНН
                'D': 80   # Ссылка
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Настройка высоты строк
            for row in range(1, len(sellers_data) + 2):
                ws.row_dimensions[row].height = 25
            
            # Применяем автофильтр
            ws.auto_filter.ref = f"A1:D{len(sellers_data) + 1}"
            
            # Замораживаем первую строку
            ws.freeze_panes = "A2"
            
            # Сохраняем файл
            wb.save(filepath)
            self.logger.info(f"Excel файл с данными продавцов сохранен: {filepath}")
            
            return filepath
            
        except Exception as e:
            self.logger.error(f"Ошибка сохранения Excel файла с продавцами: {str(e)}")
            return None
        
    def save_sellers_from_products(self, sellers_data, filename_prefix="sellers_from_products"):
        """Сохранение данных о продавцах из товаров в Excel"""
        try:
            timestamp = datetime.now().strftime("%d.%m.%Y_%H-%M-%S")
            filename = f"{self._clean_filename(filename_prefix)}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Продавцы"
            
            headers = [
                "Название продавца",
                "Название компании",
                "ИНН",
                "Ссылка на товар"
            ]
            
            # Стилизация заголовков
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self._get_border()
            
            # Заполнение данных
            for row, seller in enumerate(sellers_data, 2):
                ws.cell(row=row, column=1, value=seller.get('seller_name', ''))
                ws.cell(row=row, column=2, value=seller.get('company_name', ''))
                ws.cell(row=row, column=3, value=seller.get('inn', ''))
                ws.cell(row=row, column=4, value=seller.get('product_url', ''))
                
                # Стилизация ячеек
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self._get_border()
                    cell.alignment = Alignment(vertical="center")
                    if col == 3:  # Центрирование ИНН
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ширина колонок
            column_widths = {
                'A': 50,  # Название продавца
                'B': 60,  # Компания
                'C': 20,  # ИНН
                'D': 80   # Ссылка
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Высота строк
            for row in range(1, len(sellers_data) + 2):
                ws.row_dimensions[row].height = 25
            
            ws.auto_filter.ref = f"A1:D{len(sellers_data) + 1}"
            ws.freeze_panes = "A2"
            
            wb.save(filepath)
            self.logger.info(f"Файл продавцов сохранен: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Ошибка сохранения файла продавцов: {str(e)}")
            return None
    
    def _get_border(self):
        """Создание границ для ячеек"""
        thin_border = Side(border_style="thin", color="000000")
        return Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border
        )
    
    def _clean_filename(self, filename):
        """Очистка имени файла от недопустимых символов"""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        return filename