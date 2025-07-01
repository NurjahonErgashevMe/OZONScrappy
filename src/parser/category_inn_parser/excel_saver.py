import logging
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger('parser.excel_saver')

class ExcelSaver:
    def __init__(self, output_dir="output"):
        self.output_dir = output_dir
        
        # Создаем папку output если её нет
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def _get_border(self):
        """Создание границ для ячеек Excel"""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _apply_cell_style(self, cell, is_header=False):
        """Применение стилей к ячейке"""
        cell.border = self._get_border()
        cell.alignment = Alignment(horizontal="center" if is_header else "left", vertical="center")
        
        if is_header:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        else:
            cell.font = Font(size=11)

    def save_to_excel(self, sellers_data, category_name):
        """Сохранение данных в Excel файл"""
        try:
            # Генерируем имя файла
            timestamp = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
            safe_category_name = "".join(c for c in category_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            filename = f"{safe_category_name}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            logger.info(f"Создание Excel файла: {filename}")
            
            # Создаем рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Продавцы"
            
            # Добавляем новую колонку "Имя продавца"
            headers = ["Имя продавца", "Название компании", "ИНН", "Ссылка на товар"]
            
            # Настройка заголовков
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                self._apply_cell_style(cell, is_header=True)
            
            # Заполняем данными
            row = 2
            seller_number = 1
            
            for seller_key, data in sellers_data.items():
                # Пропускаем служебные данные
                if seller_key.startswith('_'):
                    continue
                
                seller_name = data.get('seller_name', seller_key)  # Используем распарсенное имя
                company_name = data.get('company_name', 'Не найдено')
                inn = data.get('inn', 'Не найдено')
                products = data.get('sample_products', [])
                
                if products:
                    # Для каждого товара создаем отдельную строку
                    for i, product in enumerate(products):
                        # Номер только для первого товара продавца
                        ws.cell(row=row, column=1, value=seller_name if i == 0 else "")  # Новая колонка
                        ws.cell(row=row, column=2, value=company_name if i == 0 else "")  # Сдвинуто
                        ws.cell(row=row, column=3, value=inn if i == 0 else "")           # Сдвинуто
                        ws.cell(row=row, column=4, value=product.get('url', ''))          # Сдвинуто
                        
                        # Применяем стили ко всем колонкам
                        for col in range(1, 5):
                            cell = ws.cell(row=row, column=col)
                            self._apply_cell_style(cell)
                        
                        row += 1
                else:
                    # Если нет товаров, добавляем строку с именем продавца
                    ws.cell(row=row, column=1, value=seller_name)    # Новая колонка
                    ws.cell(row=row, column=2, value=company_name)   # Сдвинуто
                    ws.cell(row=row, column=3, value=inn)            # Сдвинуто
                    ws.cell(row=row, column=4, value='')             # Сдвинуто
                    
                    for col in range(1, 5):
                        cell = ws.cell(row=row, column=col)
                        self._apply_cell_style(cell)
                    
                    row += 1
                
                seller_number += 1
            
            # Обновляем ширину колонок
            column_widths = {
                'A': 30,  # Имя продавца (новое)
                'B': 50,  # Название компании
                'C': 20,  # ИНН
                'D': 80   # Ссылка на товар
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Настройка высоты строк
            for row_num in range(1, row):
                ws.row_dimensions[row_num].height = 25
            
            # Применяем автофильтр только если есть данные
            if row > 2:
                ws.auto_filter.ref = f"A1:D{row-1}"  # Обновлено до 4 колонок
            
            # Замораживаем первую строку
            ws.freeze_panes = "A2"
            
            # Добавляем лист со статистикой
            self._create_stats_sheet(wb, sellers_data, category_name)
            
            # Сохраняем файл
            wb.save(filepath)
            logger.info(f"Excel файл успешно сохранен: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"Ошибка сохранения Excel файла: {str(e)}")
            return None

    def _create_stats_sheet(self, workbook, sellers_data, category_name):
        """Создание листа со статистикой"""
        try:
            stats_ws = workbook.create_sheet("Статистика")
            
            # Заголовок
            stats_ws.cell(row=1, column=1, value=f"Статистика парсинга: {category_name}")
            header_cell = stats_ws.cell(row=1, column=1)
            header_cell.font = Font(bold=True, size=14)
            stats_ws.merge_cells('A1:B1')
            
            # Подсчет статистики
            total_sellers = len([k for k in sellers_data.keys() if not k.startswith('_')])
            total_products = sum(len(data.get('sample_products', [])) for seller_name, data in sellers_data.items() if not seller_name.startswith('_'))
            sellers_with_inn = len([k for k, v in sellers_data.items() if not k.startswith('_') and v.get('inn') != 'Не найдено'])
            
            # Данные статистики
            stats_data = [
                ("Дата создания:", datetime.now().strftime("%d.%m.%Y %H:%M")),
                ("Всего найдено продавцов:", total_sellers),
                ("Продавцов с найденным ИНН:", sellers_with_inn),
                ("Всего товаров:", total_products),
                ("Категория:", category_name)
            ]
            
            # Заполняем статистику
            row = 3
            for label, value in stats_data:
                stats_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                stats_ws.cell(row=row, column=2, value=value)
                row += 1
            
            # Настройка ширины колонок
            stats_ws.column_dimensions['A'].width = 30
            stats_ws.column_dimensions['B'].width = 40
            
        except Exception as e:
            logger.error(f"Ошибка создания листа статистики: {str(e)}")

    def get_file_size(self, filepath):
        """Получение размера файла"""
        try:
            if os.path.exists(filepath):
                size_bytes = os.path.getsize(filepath)
                # Конвертируем в удобочитаемый формат
                if size_bytes < 1024:
                    return f"{size_bytes} B"
                elif size_bytes < 1024 * 1024:
                    return f"{size_bytes / 1024:.1f} KB"
                else:
                    return f"{size_bytes / (1024 * 1024):.1f} MB"
            return "0 B"
        except Exception as e:
            logger.error(f"Ошибка получения размера файла: {str(e)}")
            return "Unknown"