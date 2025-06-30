import logging
import asyncio
import time
import os
import re
import threading
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from aiogram.types import Message, FSInputFile
from aiogram.fsm.context import FSMContext
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.chrome.options import Options
import selenium_stealth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src.parser.seller_info_parser import SellerInfoParser
from src.bot.keyboards import main_keyboard
from src.utils import load_config

logger = logging.getLogger('parser.category_inn_parser')

class CategoryParser:
    def __init__(self):
        self.config = load_config("config.txt")
        self.seller_parser = SellerInfoParser()
        self.total_links = int(self.config.get("TOTAL_LINKS", "100"))
        self.max_idle_scrolls = int(self.config.get("MAX_IDLE_SCROLLS", "100"))
        self.scroll_delay = float(self.config.get("SCROLL_DELAY", "2.0"))
        self.load_timeout = int(self.config.get("LOAD_TIMEOUT", "30"))
        self.workers_count = 3  # Количество воркеров для парсинга ИНН
        self.output_dir = "output"
        
        # Создаем папку output если её нет
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        
    def setup_driver(self):
        """Настройка веб-драйвера с selenium-stealth"""
        options = Options()
        
        # Базовые настройки Chrome
        # options.add_argument('--headless')  # Оставляем закомментированным для отладки
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')
        options.add_argument('--start-maximized')
        options.add_argument('--disable-features=VizDisplayCompositor')
        
        # Дополнительные настройки для обхода детекции
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        
        # Настройки User-Agent
        options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        # Создание драйвера
        driver = webdriver.Chrome(options=options)
        
        # Применение stealth настроек
        selenium_stealth.stealth(
            driver,
            languages=["ru-RU", "ru"],
            vendor="Google Inc.",
            platform="Win32",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=True,
        )
        
        # Дополнительные скрипты для маскировки
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        driver.execute_cdp_cmd('Network.setUserAgentOverride', {
            "userAgent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        })
        
        return driver

    def _simulate_human_behavior(self, driver):
        """Имитация человеческого поведения для обхода детекции"""
        try:
            # Случайный скролл
            driver.execute_script("window.scrollTo(0, Math.floor(Math.random() * 500));")
            time.sleep(1)
            
            # Движение мыши (имитация через JavaScript)
            driver.execute_script("""
                var event = new MouseEvent('mousemove', {
                    'view': window,
                    'bubbles': true,
                    'cancelable': true,
                    'clientX': Math.random() * window.innerWidth,
                    'clientY': Math.random() * window.innerHeight
                });
                document.dispatchEvent(event);
            """)
            time.sleep(0.5)
            
        except Exception as e:
            logger.debug(f"Ошибка при имитации поведения: {str(e)}")

    def get_category_name(self, url):
        """Извлечение имени категории из URL"""
        try:
            # Извлекаем последнюю часть URL после последнего слеша
            category_part = url.rstrip('/').split('/')[-1]
            
            # Убираем параметры запроса
            if '?' in category_part:
                category_part = category_part.split('?')[0]
            
            # Заменяем дефисы на подчеркивания для имени файла
            category_name = category_part.replace('-', '_')
            
            # Если не удалось извлечь осмысленное имя, используем дефолтное
            if not category_name or len(category_name) < 3:
                category_name = "category"
                
            return category_name
        except Exception as e:
            logger.warning(f"Ошибка извлечения имени категории: {e}")
            return "category"

    def collect_product_links(self, category_url):
        """Сбор ссылок на товары из категории"""
        logger.info(f"Начинаем сбор ссылок из категории: {category_url}")
        
        driver = None
        unique_links = set()
        ordered_links = []
        idle_scrolls = 0
        
        try:
            driver = self.setup_driver()
            
            # Загружаем страницу категории
            logger.info(f"Загрузка страницы: {category_url}")
            driver.get(category_url)
            
            # Ожидаем появления товаров
            try:
                WebDriverWait(driver, self.load_timeout).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".tile-root"))
                )
                logger.info("Страница успешно загружена")
            except TimeoutException:
                logger.error("Не удалось дождаться загрузки товаров")
                return []
            
            # Первоначальный сбор ссылок
            current_links = self._extract_product_links(driver)
            if current_links:
                unique_links.update(current_links)
                ordered_links.extend(current_links)
                logger.info(f"Начальное количество ссылок: {len(ordered_links)}")
            
            # Основной цикл сбора ссылок через скролл
            while len(ordered_links) < self.total_links and idle_scrolls < self.max_idle_scrolls:
                # Плавный скролл страницы
                self._scroll_page(driver)
                
                # Имитируем человеческое поведение
                self._simulate_human_behavior(driver)
                
                # Собираем новые ссылки
                current_links = self._extract_product_links(driver)
                new_links = current_links - unique_links
                
                if new_links:
                    logger.info(f"Найдено новых ссылок: {len(new_links)}")
                    unique_links.update(new_links)
                    ordered_links.extend(new_links)
                    idle_scrolls = 0  # Сбрасываем счетчик
                    
                    current_count = len(ordered_links)
                    logger.info(f"Всего ссылок: {min(current_count, self.total_links)}/{self.total_links}")
                else:
                    idle_scrolls += 1
                    logger.debug(f"Нет новых ссылок, idle_scrolls: {idle_scrolls}")
                
                # Проверяем конец страницы
                if self._is_page_end_reached(driver):
                    logger.info("Достигнут конец страницы")
                    break
                
                time.sleep(0.5)
            
            # Ограничиваем количество ссылок
            final_links = ordered_links[:self.total_links]
            
            # Сохраняем ссылки в файл
            category_name = self.get_category_name(category_url)
            self._save_links_to_file(final_links, category_name)
            
            logger.info(f"Сбор ссылок завершен. Собрано: {len(final_links)} ссылок")
            return final_links
            
        except Exception as e:
            logger.error(f"Ошибка при сборе ссылок: {str(e)}")
            return []
        finally:
            if driver:
                try:
                    driver.quit()
                    logger.info("Драйвер для сбора ссылок закрыт")
                except Exception as e:
                    logger.warning(f"Ошибка при закрытии драйвера: {str(e)}")

    def _extract_product_links(self, driver):
        """Извлечение ссылок на товары со страницы"""
        try:
            # Используем селектор из link_parser.py
            elements = driver.find_elements(By.CSS_SELECTOR, ".tile-root a.tile-clickable-element")
            links = set()
            
            for element in elements:
                href = element.get_attribute("href")
                if href and "/product/" in href:
                    # Очищаем URL от лишних параметров, оставляя только основную часть
                    clean_href = href.split('?')[0] if '?' in href else href
                    if clean_href.startswith("https://www.ozon.ru/product/") or clean_href.startswith("/product/"):
                        # Преобразуем относительный URL в абсолютный
                        if clean_href.startswith("/product/"):
                            clean_href = "https://www.ozon.ru" + clean_href
                        links.add(clean_href)
            
            return links
        except Exception as e:
            logger.warning(f"Ошибка извлечения ссылок: {str(e)}")
            return set()

    def _scroll_page(self, driver):
        """Плавный скролл страницы"""
        try:
            current_scroll_position = driver.execute_script("return window.pageYOffset")
            scroll_height = driver.execute_script("return document.body.scrollHeight")
            
            # Плавный скролл порциями по 300px
            for i in range(current_scroll_position, scroll_height, 300):
                driver.execute_script(f"window.scrollTo(0, {i});")
                time.sleep(0.05)
            
            time.sleep(self.scroll_delay)
            
        except Exception as e:
            logger.warning(f"Ошибка при скролле: {str(e)}")

    def _is_page_end_reached(self, driver):
        """Проверка достижения конца страницы"""
        try:
            total_height = driver.execute_script("return document.body.scrollHeight")
            current_position = driver.execute_script("return window.pageYOffset + window.innerHeight")
            return (total_height - current_position) < 100
        except Exception as e:
            logger.warning(f"Ошибка проверки конца страницы: {str(e)}")
            return False

    def _save_links_to_file(self, links, category_name):
        """Сохранение ссылок в файл"""
        try:
            # Создаем директорию output если её нет
            os.makedirs("output", exist_ok=True)
            
            filename = f"output/links_{category_name}.txt"
            with open(filename, "w", encoding="utf-8") as f:
                f.write("\n".join(links))
            
            logger.info(f"Ссылки сохранены в файл: {os.path.abspath(filename)}")
        except Exception as e:
            logger.error(f"Ошибка сохранения ссылок: {str(e)}")

    def _get_border(self):
        """Создание границ для ячеек Excel"""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def save_to_excel(self, sellers_data, category_name):
        """Сохранение данных в Excel файл"""
        try:
            timestamp = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
            filename = f"{category_name}_sellers_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            # Создаем рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Продавцы"
            
            # Заголовки
            headers = ["Название компании", "ИНН", "Ссылка на товар"]
            
            # Настройка заголовков
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self._get_border()
            
            # Заполняем данными
            row = 2
            for seller_name, data in sellers_data.items():
                company_name = data.get('company_name', 'Не найдено')
                inn = data.get('inn', 'Не найдено')
                
                # Для каждого товара создаем отдельную строку
                for product in data.get('sample_products', []):
                    ws.cell(row=row, column=1, value=company_name)
                    ws.cell(row=row, column=2, value=inn)
                    ws.cell(row=row, column=3, value=product.get('url', ''))
                    
                    # Применяем стили к ячейкам
                    for col in range(1, 4):
                        cell = ws.cell(row=row, column=col)
                        cell.border = self._get_border()
                        cell.alignment = Alignment(vertical="center")
                    
                    row += 1
                
                # Если нет товаров, все равно добавляем строку с компанией
                if not data.get('sample_products'):
                    ws.cell(row=row, column=1, value=company_name)
                    ws.cell(row=row, column=2, value=inn)
                    ws.cell(row=row, column=3, value='')
                    
                    for col in range(1, 4):
                        cell = ws.cell(row=row, column=col)
                        cell.border = self._get_border()
                        cell.alignment = Alignment(vertical="center")
                    
                    row += 1
            
            # Настройка ширины колонок
            ws.column_dimensions['A'].width = 75  # Название компании
            ws.column_dimensions['B'].width = 30  # ИНН
            ws.column_dimensions['C'].width = 75  # Ссылка на товар
            
            # Настройка высоты строк
            for row_num in range(1, row):
                ws.row_dimensions[row_num].height = 25
            
            # Применяем автофильтр
            ws.auto_filter.ref = f"A1:C{row-1}"
            
            # Замораживаем первую строку
            ws.freeze_panes = "A2"
            
            # Сохраняем файл
            wb.save(filepath)
            logger.info(f"Excel файл сохранен: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"Ошибка сохранения Excel файла: {str(e)}")
            return None

    def save_inn_to_txt(self, sellers_data, category_name):
        """Сохранение только ИНН в текстовый файл"""
        try:
            timestamp = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
            filename = f"{category_name}_inn_{timestamp}.txt"
            filepath = os.path.join(self.output_dir, filename)
            
            # Собираем уникальные ИНН
            inn_list = []
            for data in sellers_data.values():
                inn = data.get('inn', '')
                if inn and inn != 'Не найдено' and inn not in inn_list:
                    inn_list.append(inn)
            
            # Сохраняем в файл
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write('\n'.join(inn_list))
            
            logger.info(f"Файл с ИНН сохранен: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Ошибка сохранения файла с ИНН: {str(e)}")
            return None

    async def send_file_to_user(self, bot, chat_id, filepath, caption):
        """Отправка файла пользователю"""
        try:
            if os.path.exists(filepath):
                file = FSInputFile(filepath)
                await bot.send_document(chat_id, file, caption=caption)
                logger.info(f"Файл отправлен пользователю: {filepath}")
            else:
                logger.error(f"Файл не найден: {filepath}")
        except Exception as e:
            logger.error(f"Ошибка отправки файла: {str(e)}")

    def parse_sellers_from_links(self, links):
        """Парсинг ИНН продавцов из списка ссылок с использованием 3 воркеров"""
        logger.info(f"Начинаем парсинг ИНН из {len(links)} ссылок с {self.workers_count} воркерами")
        
        all_sellers = {}
        sellers_lock = threading.Lock()
        
        def worker_parse_links(worker_id, worker_links):
            """Функция воркера для парсинга ссылок"""
            driver = None
            try:
                driver = self.setup_driver()
                logger.info(f"Воркер {worker_id}: Обрабатывает {len(worker_links)} ссылок")
                
                for i, link in enumerate(worker_links, 1):
                    try:
                        logger.info(f"Воркер {worker_id}: Обрабатывает ссылку {i}/{len(worker_links)}: {link}")
                        
                        # Переходим на страницу товара
                        driver.get(link)
                        time.sleep(4)
                        
                        # Имитируем человеческое поведение
                        self._simulate_human_behavior(driver)
                        
                        # Получаем информацию о продавце
                        seller_info = self.seller_parser.get_seller_details(driver)
                        
                        if seller_info and seller_info.get('company_name') != 'Не найдено':
                            seller_key = seller_info.get('company_name', 'Неизвестно')
                            
                            # Потокобезопасное обновление общего словаря
                            with sellers_lock:
                                if seller_key not in all_sellers:
                                    all_sellers[seller_key] = {
                                        'company_name': seller_info.get('company_name', 'Не найдено'),
                                        'inn': seller_info.get('inn', 'Не найдено'),
                                        'products_count': 0,
                                        'sample_products': []
                                    }
                                
                                all_sellers[seller_key]['products_count'] += 1
                                
                                # Добавляем примеры товаров (максимум 3)
                                if len(all_sellers[seller_key]['sample_products']) < 3:
                                    try:
                                        product_title = driver.find_element(By.CSS_SELECTOR, 'h1').text[:100]
                                        all_sellers[seller_key]['sample_products'].append({
                                            'title': product_title,
                                            'url': link
                                        })
                                    except:
                                        pass
                            
                            logger.info(f"Воркер {worker_id}: ✓ Добавлен продавец: {seller_key}")
                        else:
                            logger.warning(f"Воркер {worker_id}: Не удалось получить информацию о продавце")
                            
                    except Exception as e:
                        logger.error(f"Воркер {worker_id}: Ошибка при обработке ссылки {i}: {str(e)}")
                        continue
                        
            except Exception as e:
                logger.error(f"Воркер {worker_id}: Критическая ошибка: {str(e)}")
            finally:
                if driver:
                    try:
                        driver.quit()
                        logger.info(f"Воркер {worker_id}: Драйвер закрыт")
                    except Exception as e:
                        logger.warning(f"Воркер {worker_id}: Ошибка при закрытии драйвера: {str(e)}")

        # Разделяем ссылки между воркерами
        worker_links = [[] for _ in range(self.workers_count)]
        for i, link in enumerate(links):
            worker_id = i % self.workers_count
            worker_links[worker_id].append(link)
        
        # Логируем распределение ссылок
        for i, links_list in enumerate(worker_links):
            logger.info(f"Воркер {i+1}: назначено {len(links_list)} ссылок")
        
        # Запускаем воркеры
        with ThreadPoolExecutor(max_workers=self.workers_count) as executor:
            futures = []
            for i in range(self.workers_count):
                if worker_links[i]:  # Запускаем только если есть ссылки для обработки
                    future = executor.submit(worker_parse_links, i+1, worker_links[i])
                    futures.append(future)
            
            # Ждем завершения всех воркеров
            for future in futures:
                try:
                    future.result()
                except Exception as e:
                    logger.error(f"Ошибка в воркере: {str(e)}")
        
        logger.info(f"Парсинг ИНН завершен. Найдено {len(all_sellers)} уникальных продавцов")
        return all_sellers

    def parse_category(self, category_url):
        """Основной метод парсинга категории"""
        logger.info(f"Начинаем полный парсинг категории: {category_url}")
        
        try:
            # Этап 1: Сбор ссылок на товары
            logger.info("=== ЭТАП 1: Сбор ссылок на товары ===")
            product_links = self.collect_product_links(category_url)
            
            if not product_links:
                logger.error("Не удалось собрать ссылки на товары")
                return {}
            
            logger.info(f"Собрано {len(product_links)} ссылок на товары")
            
            # Этап 2: Парсинг ИНН продавцов
            logger.info("=== ЭТАП 2: Парсинг ИНН продавцов ===")
            sellers_data = self.parse_sellers_from_links(product_links)
            
            if sellers_data:
                category_name = self.get_category_name(category_url)
                
                # Этап 3: Создание Excel файла
                logger.info("=== ЭТАП 3: Создание Excel файла ===")
                excel_filepath = self.save_to_excel(sellers_data, category_name)
                
                # Этап 4: Создание TXT файла с ИНН
                logger.info("=== ЭТАП 4: Создание файла с ИНН ===")
                txt_filepath = self.save_inn_to_txt(sellers_data, category_name)
                
                # Сохраняем пути к файлам для последующей отправки
                sellers_data['_files'] = {
                    'excel': excel_filepath,
                    'txt': txt_filepath,
                    'category_name': category_name
                }
            
            return sellers_data
            
        except Exception as e:
            logger.error(f"Критическая ошибка при парсинге категории: {str(e)}")
            return {}

    async def send_files_to_user(self, bot, chat_id, sellers_data):
        """Отправка файлов пользователю после парсинга"""
        try:
            if '_files' not in sellers_data:
                return
                
            files_info = sellers_data['_files']
            category_name = files_info['category_name']
            
            # Отправляем Excel файл
            if files_info['excel']:
                await self.send_file_to_user(
                    bot, chat_id, files_info['excel'], 
                    f"📊 Excel файл с данными продавцов категории {category_name}"
                )
            
            # Отправляем TXT файл
            if files_info['txt']:
                await self.send_file_to_user(
                    bot, chat_id, files_info['txt'], 
                    f"📋 Текстовый файл с ИНН категории {category_name}"
                )
                
            # Удаляем служебную информацию
            del sellers_data['_files']
            
        except Exception as e:
            logger.error(f"Ошибка при отправке файлов: {str(e)}")

    def format_results(self, sellers_data):
        """Упрощенное форматирование результатов для отправки"""
        if not sellers_data:
            return "❌ Продавцы не найдены"
            
        result = f"📊 **Результаты парсинга категории**\n\n"
        result += f"🔍 Найдено продавцов: **{len(sellers_data)}**\n\n"
        
        # Сортируем по количеству товаров
        sorted_sellers = sorted(sellers_data.items(), key=lambda x: x[1]['products_count'], reverse=True)
        
        for i, (seller_name, data) in enumerate(sorted_sellers, 1):
            result += f"**{i}. {data['company_name']}**\n"
            result += f"   📋 ИНН: `{data['inn']}`\n"
            result += f"   📦 Товаров найдено: {data['products_count']}\n\n"
        
        result += f"📁 Файлы с подробными данными отправлены отдельно"
            
        return result